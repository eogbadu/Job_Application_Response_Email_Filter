from pathlib import Path
import openpyxl

COLUMNS = ["Company", "Job Title", "Date Applied", "Notes"]
_SKIP_WORDS = {"the", "and", "inc", "llc", "ltd", "corp", "co", "of", "for", "a", "an"}


def load_applied_companies(filepath: str) -> list[str]:
    path = Path(filepath)
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    companies = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row and row[0]:
            name = str(row[0]).strip()
            if name:
                companies.append(name)
    wb.close()
    return companies


def create_template(filepath: str) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Applied Jobs"
    ws.append(COLUMNS)
    # Lock column widths for readability
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 35
    ws.column_dimensions["C"].width = 15
    ws.column_dimensions["D"].width = 40
    wb.save(filepath)


def company_matches(company: str, sender: str, subject: str) -> bool:
    """Return True if the company name appears in the sender address or subject line."""
    company_lower = company.lower()
    search = f"{sender.lower()} {subject.lower()}"

    # Full name match
    if company_lower in search:
        return True

    # Significant-word match: any word >3 chars not in the skip list
    words = [
        w for w in company_lower.split()
        if len(w) > 3 and w not in _SKIP_WORDS
    ]
    return any(w in search for w in words)
